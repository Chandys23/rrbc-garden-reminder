#!/usr/bin/env python3
"""
Bulk Import Gardeners from Excel File
Reads Schedule.xls and inserts all gardeners into the database
"""

import sys
from pathlib import Path
import sqlite3
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from database import get_connection, init_db

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class BulkImporter:
    def __init__(self):
        self.excel_file = Path(__file__).parent / "Schedule.xlsx"
        self.imported_count = 0
        self.error_count = 0
        self.errors = []
        
    def validate_date(self, date_str):
        """Validate date format (YYYY-MM-DD)"""
        try:
            from datetime import datetime as dt
            # If it's already a datetime object, just check it
            if isinstance(date_str, dt):
                return True
            # Otherwise try to parse as string
            dt.strptime(str(date_str).strip(), '%Y-%m-%d')
            return True
        except (ValueError, AttributeError):
            return False
    
    def validate_email(self, email):
        """Basic email validation"""
        email = str(email).strip()
        return '@' in email and '.' in email
    
    def read_excel(self):
        """Read Excel file and return list of dictionaries"""
        if not self.excel_file.exists():
            print(f"❌ File not found: {self.excel_file}")
            return None
        
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Get header row (first row)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value.lower() if cell.value else "")
            
            # Validate headers
            required_headers = ['date', 'task', 'name', 'email', 'phone']
            if not all(h in headers for h in required_headers):
                print("❌ Excel file missing required columns!")
                print(f"   Required: {required_headers}")
                print(f"   Found: {headers}")
                return None
            
            # Read data rows
            data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue
                
                record = {}
                for col_idx, header in enumerate(headers):
                    record[header] = row[col_idx]
                
                data.append((row_idx, record))
            
            return data
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            return None
    
    def validate_record(self, row_idx, record):
        """Validate a single record"""
        errors = []
        
        # Check required fields
        if not record.get('date'):
            errors.append("Missing date")
        elif not self.validate_date(record['date']):
            errors.append(f"Invalid date format: {record['date']} (use YYYY-MM-DD)")
        
        if not record.get('task'):
            errors.append("Missing task")
        
        if not record.get('name'):
            errors.append("Missing name")
        
        if not record.get('email'):
            errors.append("Missing email")
        elif not self.validate_email(record['email']):
            errors.append(f"Invalid email: {record['email']}")
        
        if not record.get('phone'):
            errors.append("Missing phone")
        
        return errors
    
    def insert_gardener(self, record):
        """Insert single gardener into database"""
        try:
            # Convert date if it's a datetime object
            date_value = record['date']
            if isinstance(date_value, datetime):
                date_value = date_value.strftime('%Y-%m-%d')
            else:
                date_value = str(date_value).strip()
            
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO gardeners (date, task, name, email, mobile)
                VALUES (?, ?, ?, ?, ?)
            """, (
                date_value,
                str(record['task']).strip(),
                str(record['name']).strip(),
                str(record['email']).strip(),
                str(record['phone']).strip()
            ))
            
            conn.commit()
            conn.close()
            return True
            
        except Exception as e:
            print(f"   Error inserting: {str(e)}")
            return False
    
    def import_from_excel(self):
        """Main import function"""
        print("=" * 60)
        print("🌿 RRBC Garden Care - Bulk Import Gardeners")
        print("=" * 60)
        print()
        
        # Check if file exists
        if not self.excel_file.exists():
            print(f"❌ Schedule.xls not found at: {self.excel_file}")
            return False
        
        print(f"📁 Reading: {self.excel_file}")
        
        # Read Excel
        records = self.read_excel()
        if records is None:
            return False
        
        if not records:
            print("⚠️  No data found in Excel file")
            return False
        
        print(f"📊 Found {len(records)} gardeners to import")
        print()
        
        # Initialize database
        try:
            init_db()
        except Exception as e:
            print(f"⚠️  Database init warning: {str(e)}")
        
        print("Processing records:")
        print("-" * 60)
        
        # Validate and import each record
        for row_idx, record in records:
            print(f"\nRow {row_idx}: {record.get('name', 'Unknown')}")
            
            # Validate
            errors = self.validate_record(row_idx, record)
            
            if errors:
                self.error_count += 1
                print(f"  ❌ Validation failed:")
                for error in errors:
                    print(f"     - {error}")
                self.errors.append((row_idx, record.get('name', '?'), errors))
                continue
            
            # Insert
            if self.insert_gardener(record):
                self.imported_count += 1
                print(f"  ✅ Imported: {record['name']} | {record['date']} | {record['task']}")
            else:
                self.error_count += 1
                print(f"  ❌ Failed to insert")
                self.errors.append((row_idx, record.get('name', '?'), ["Database insert failed"]))
        
        # Summary
        print()
        print("=" * 60)
        print("📊 IMPORT SUMMARY")
        print("=" * 60)
        print(f"✅ Successfully imported: {self.imported_count} gardeners")
        print(f"❌ Failed: {self.error_count} gardeners")
        print(f"📈 Total processed: {self.imported_count + self.error_count}")
        print()
        
        if self.errors:
            print("⚠️  ERRORS:")
            for row, name, errs in self.errors:
                print(f"\n  Row {row} - {name}:")
                for err in errs:
                    print(f"    • {err}")
        
        print()
        print("=" * 60)
        
        return self.error_count == 0


if __name__ == "__main__":
    importer = BulkImporter()
    success = importer.import_from_excel()
    sys.exit(0 if success else 1)
